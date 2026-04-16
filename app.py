from flask import Flask, render_template, jsonify, request, send_file
from datetime import datetime
import json
import os
from openpyxl import load_workbook

app = Flask(__name__)

TASKS_FILE = os.path.join(os.path.dirname(__file__), 'tasks.json')

def load_tasks():
    """从 JSON 文件加载任务"""
    if not os.path.exists(TASKS_FILE):
        return {"tasks": [], "last_id": 0}
    with open(TASKS_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_tasks(data):
    """保存任务到 JSON 文件"""
    with open(TASKS_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def get_next_id():
    """获取下一个任务 ID"""
    data = load_tasks()
    data['last_id'] += 1
    save_tasks(data)
    return data['last_id']

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/tasks')
def tasks_page():
    return render_template('index.html')

@app.route('/daily')
def daily_page():
    return render_template('index.html')

@app.route('/report')
def report_page():
    return render_template('index.html')

@app.route('/record')
def record_page():
    return render_template('record.html')

@app.route('/task/<int:task_id>')
def task_detail_page(task_id):
    return render_template('index.html')

# Task CRUD API
@app.route('/api/tasks', methods=['GET'])
def get_tasks():
    """获取所有任务"""
    data = load_tasks()
    return jsonify({'success': True, 'data': data['tasks']})

@app.route('/api/tasks', methods=['POST'])
def create_task():
    """创建新任务"""
    task_data = request.get_json()
    required_fields = ['name', 'category', 'plan_start', 'plan_end']
    for field in required_fields:
        if not task_data.get(field):
            return jsonify({'success': False, 'error': f'缺少必填字段: {field}'}), 400

    now = datetime.now().strftime('%Y-%m-%d')
    task = {
        'id': get_next_id(),
        'name': task_data['name'],
        'category': task_data['category'],
        'priority': task_data.get('priority', '中'),
        'status': task_data.get('status', '未开始'),
        'plan_start': task_data['plan_start'],
        'plan_end': task_data['plan_end'],
        'progress_note': task_data.get('progress_note', ''),
        'pause_reason': task_data.get('pause_reason', ''),
        'sort_order': task_data.get('sort_order', 0),
        'created_at': now,
        'updated_at': now
    }

    data = load_tasks()
    data['tasks'].append(task)
    save_tasks(data)

    return jsonify({'success': True, 'data': task}), 201

@app.route('/api/tasks/<int:task_id>', methods=['GET'])
def get_task(task_id):
    """获取单个任务"""
    data = load_tasks()
    task = next((t for t in data['tasks'] if t['id'] == task_id), None)
    if not task:
        return jsonify({'success': False, 'error': '任务不存在'}), 404
    return jsonify({'success': True, 'data': task})

@app.route('/api/tasks/<int:task_id>', methods=['PUT'])
def update_task(task_id):
    """更新任务"""
    task_data = request.get_json()
    data = load_tasks()
    task_index = next((i for i, t in enumerate(data['tasks']) if t['id'] == task_id), None)

    if task_index is None:
        return jsonify({'success': False, 'error': '任务不存在'}), 404

    task = data['tasks'][task_index]
    task['name'] = task_data.get('name', task['name'])
    task['category'] = task_data.get('category', task['category'])
    task['priority'] = task_data.get('priority', task['priority'])
    task['status'] = task_data.get('status', task['status'])
    task['plan_start'] = task_data.get('plan_start', task['plan_start'])
    task['plan_end'] = task_data.get('plan_end', task['plan_end'])
    task['progress_note'] = task_data.get('progress_note', task['progress_note'])
    task['pause_reason'] = task_data.get('pause_reason', task['pause_reason'])
    task['sort_order'] = task_data.get('sort_order', task['sort_order'])
    task['updated_at'] = datetime.now().strftime('%Y-%m-%d')

    save_tasks(data)
    return jsonify({'success': True, 'data': task})

@app.route('/api/tasks/<int:task_id>', methods=['DELETE'])
def delete_task(task_id):
    """删除任务"""
    data = load_tasks()
    task_index = next((i for i, t in enumerate(data['tasks']) if t['id'] == task_id), None)

    if task_index is None:
        return jsonify({'success': False, 'error': '任务不存在'}), 404

    data['tasks'].pop(task_index)
    save_tasks(data)
    return jsonify({'success': True, 'message': '删除成功'})

@app.route('/api/tasks/<int:task_id>/status', methods=['PUT'])
def update_task_status(task_id):
    """更新任务状态"""
    task_data = request.get_json()
    data = load_tasks()
    task_index = next((i for i, t in enumerate(data['tasks']) if t['id'] == task_id), None)

    if task_index is None:
        return jsonify({'success': False, 'error': '任务不存在'}), 404

    task = data['tasks'][task_index]
    task['status'] = task_data.get('status', task['status'])
    task['progress_note'] = task_data.get('progress_note', task.get('progress_note', ''))
    task['pause_reason'] = task_data.get('pause_reason', task.get('pause_reason', ''))
    task['updated_at'] = datetime.now().strftime('%Y-%m-%d')

    # 记录完成时间
    if task['status'] == '已完成':
        task['completed_at'] = datetime.now().strftime('%Y-%m-%d')
    else:
        task['completed_at'] = task.get('completed_at', '')

    save_tasks(data)
    return jsonify({'success': True, 'data': task})

@app.route('/api/daily', methods=['GET'])
def get_daily_tasks():
    """按日期查询任务"""
    date = request.args.get('date')

    if not date:
        date = datetime.now().strftime('%Y-%m-%d')

    data = load_tasks()
    tasks = [t for t in data['tasks'] if t['plan_start'] <= date <= t['plan_end']]
    tasks.sort(key=lambda x: (x['category'], x['sort_order'], x['id']))

    return jsonify({'success': True, 'data': tasks})

@app.route('/api/report/preview', methods=['GET'])
def get_report_preview():
    """获取周报预览数据"""
    week_start = request.args.get('start_date')
    week_end = request.args.get('end_date')

    if not week_start or not week_end:
        return jsonify({'success': False, 'error': '缺少日期参数'}), 400

    data = load_tasks()

    # 本周计划（plan_start <= week_end AND plan_end >= week_start 的任务）
    week_plan_tasks = [t for t in data['tasks']
                       if t['plan_start'] <= week_end and t['plan_end'] >= week_start]

    # 本周完成（plan_end 在本周内且状态为已完成）
    completed_tasks = [t for t in data['tasks']
                      if t['status'] == '已完成' and week_start <= t['plan_end'] <= week_end]

    categories = ['教学', '竞赛', '就业', '科研', '项目', '职能组', '院校支撑']

    def categorize(task_list):
        result = {cat: [] for cat in categories}
        for task in task_list:
            cat = task.get('category')
            if cat in result:
                result[cat].append({
                    'id': task.get('id'),
                    'name': task.get('name', ''),
                    'progress_note': task.get('progress_note', ''),
                    'status': task.get('status', '')
                })
        return result

    return jsonify({
        'success': True,
        'data': {
            'week_plan': categorize(week_plan_tasks),
            'completed': categorize(completed_tasks)
        }
    })

@app.route('/api/report/generate', methods=['POST'])
def generate_report():
    """使用模板生成周报Excel"""
    task_data = request.get_json()
    week_start = task_data.get('start_date')
    week_end = task_data.get('end_date')

    # 前端传来的可编辑内容
    week_plan_content = task_data.get('week_plan_content', {})  # 本周计划（已编辑）
    completed_content = task_data.get('completed_content', {})  # 本周完成情况（已编辑）
    next_week_plan = task_data.get('next_week_plan', {})  # 下周计划
    coordinated_items = task_data.get('coordinated_items', '')

    if not week_start or not week_end:
        return jsonify({'success': False, 'error': '缺少日期参数'}), 400

    data = load_tasks()

    # 本周完成
    completed_tasks = [t for t in data['tasks']
                       if t['status'] == '已完成' and week_start <= t['plan_end'] <= week_end]

    # 本周计划
    week_plan_tasks = [t for t in data['tasks']
                       if t['plan_start'] <= week_end and t['plan_end'] >= week_start]

    categories = ['教学', '竞赛', '就业', '科研', '项目', '职能组', '院校支撑']

    categorized_completed = {cat: [] for cat in categories}
    categorized_plan = {cat: [] for cat in categories}

    for task in completed_tasks:
        cat = task.get('category')
        if cat in categorized_completed:
            categorized_completed[cat].append({
                'name': task.get('name', ''),
                'progress_note': task.get('progress_note', '')
            })

    for task in week_plan_tasks:
        cat = task.get('category')
        if cat in categorized_plan:
            categorized_plan[cat].append(task.get('name', ''))

    # 加载模板文件
    template_path = os.path.join(os.path.dirname(__file__), '周报模版-新.xlsx')
    wb = load_workbook(template_path)
    ws = wb.active

    # 填充基本信息
    ws['B2'] = '王弯弯'
    ws['D2'] = f'{week_start} 至 {week_end}'
    ws['B3'] = '河工大'
    ws['D3'] = datetime.now().strftime('%Y-%m-%d')

    # 填充分类数据 (A5-A11 对应 教学-院校支撑)
    for i, cat in enumerate(categories, start=5):
        # 本周计划（B列）- 使用前端编辑内容
        if week_plan_content and isinstance(week_plan_content, dict) and week_plan_content.get(cat):
            ws[f'B{i}'] = week_plan_content[cat]
        else:
            plan_tasks = categorized_plan.get(cat, [])
            if plan_tasks:
                plan_texts = [f"{idx}.{name}" for idx, name in enumerate(plan_tasks, 1)]
                ws[f'B{i}'] = '\n'.join(plan_texts)

        # 本周完成情况（C列）- 使用前端编辑内容
        if completed_content and isinstance(completed_content, dict) and completed_content.get(cat):
            ws[f'C{i}'] = completed_content[cat]
        else:
            tasks_in_cat = categorized_completed.get(cat, [])
            if tasks_in_cat:
                task_texts = []
                for idx, t in enumerate(tasks_in_cat, 1):
                    name = t['name']
                    note = t.get('progress_note', '')
                    if note:
                        task_texts.append(f"{idx}.完成{name}：{note}")
                    else:
                        task_texts.append(f"{idx}.完成{name}")
                ws[f'C{i}'] = '\n'.join(task_texts)

        # 下周工作计划（D列）- 使用 next_week_plan
        if next_week_plan and isinstance(next_week_plan, dict) and next_week_plan.get(cat):
            ws[f'D{i}'] = next_week_plan[cat]

    # 待协调事项（E5-E11，按分类）
    if coordinated_items:
        if isinstance(coordinated_items, dict):
            for i, cat in enumerate(categories, start=5):
                if coordinated_items.get(cat):
                    ws[f'E{i}'] = coordinated_items[cat]
        else:
            ws['E5'] = coordinated_items

    # 保存到临时文件
    filename = f"{week_start}-{week_end}-周报.xlsx"
    import tempfile
    temp_dir = tempfile.gettempdir()
    filepath = os.path.join(temp_dir, filename)
    wb.save(filepath)

    return send_file(filepath, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    app.run(debug=True, port=5000)
